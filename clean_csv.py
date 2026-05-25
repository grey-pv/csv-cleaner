"""
CSV Cleaner Tool
----------------
Cleans messy CSV files and exports a cleaned version + an Excel report
showing exactly what was fixed.

Usage:
    python clean_csv.py data.csv
    python clean_csv.py data.csv --output clean_data.csv
    python clean_csv.py *.csv
    python clean_csv.py ./folder
"""

import sys
import argparse
import re
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Toolkit visual style ───────────────────────────────────────────────────────
HEADER_BG   = "1F4E79"
HEADER_FG   = "FFFFFF"
ALT_ROW_BG  = "D6E4F0"
BORDER_CLR  = "A9C4D8"
GREEN_BG    = "E2EFDA"
RED_BG      = "FCE4D6"


# ── Cleaning functions ─────────────────────────────────────────────────────────

def fix_encoding(path: Path) -> tuple[pd.DataFrame, str]:
    """Try reading the CSV with different encodings until one works."""
    encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252", "iso-8859-1"]
    for enc in encodings:
        try:
            df = pd.read_csv(path, encoding=enc)
            return df, enc
        except (UnicodeDecodeError, Exception):
            continue
    raise ValueError(f"Could not read {path.name} with any known encoding.")


def clean_column_names(df: pd.DataFrame) -> tuple[pd.DataFrame, list]:
    """
    Standardizes column names:
    - Strip whitespace
    - Lowercase
    - Replace spaces and special chars with underscores
    """
    original = list(df.columns)
    df.columns = (
        df.columns
        .str.strip()
        .str.lower()
        .str.replace(r"[^\w]", "_", regex=True)
        .str.replace(r"_+", "_", regex=True)
        .str.strip("_")
    )
    changed = [(o, n) for o, n in zip(original, df.columns) if o != n]
    return df, changed


def remove_empty_rows(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Remove rows where ALL values are NaN or empty string."""
    before = len(df)
    df = df.replace("", pd.NA)
    df = df.dropna(how="all")
    df = df.reset_index(drop=True)
    return df, before - len(df)


def remove_duplicates(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Remove fully duplicate rows."""
    before = len(df)
    df = df.drop_duplicates()
    df = df.reset_index(drop=True)
    return df, before - len(df)


def trim_whitespace(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Strip leading/trailing whitespace from all string cells."""
    count = 0
    for col in df.select_dtypes(include=["object", "str"]).columns:
        original = df[col].copy()
        df[col] = df[col].str.strip()
        count += (original != df[col]).sum()
    return df, int(count)


def standardize_dates(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """
    Detects columns likely to contain dates and standardizes to YYYY-MM-DD.
    Only touches columns with 'date', 'dt', 'time', 'day' in the name.
    """
    count = 0
    date_keywords = ["date", "dt", "time", "day", "month", "year", "data"]
    date_formats  = [
        "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y",
        "%d/%m/%y", "%m/%d/%y", "%Y/%m/%d", "%d.%m.%Y",
    ]

    for col in df.columns:
        if not any(kw in col.lower() for kw in date_keywords):
            continue
        if df[col].dtype != object:
            continue

        converted = 0
        new_col = df[col].copy()

        for idx, val in df[col].items():
            if pd.isna(val) or str(val).strip() == "":
                continue
            for fmt in date_formats:
                try:
                    parsed = datetime.strptime(str(val).strip(), fmt)
                    new_col.at[idx] = parsed.strftime("%Y-%m-%d")
                    converted += 1
                    break
                except ValueError:
                    continue

        if converted > 0:
            df[col] = new_col
            count += converted

    return df, count


# ── Excel report ───────────────────────────────────────────────────────────────

def style_sheet(ws, df: pd.DataFrame, header_color: str = HEADER_BG):
    """Applies standard toolkit formatting to a worksheet."""
    thin   = Side(style="thin", color=BORDER_CLR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = Font(bold=True, color=HEADER_FG, name="Calibri", size=11)
        cell.fill      = PatternFill("solid", fgColor=header_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        bg = ALT_ROW_BG if row_idx % 2 == 0 else "FFFFFF"
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = border
            cell.font      = Font(name="Calibri", size=10)

    for col_idx, col in enumerate(df.columns, start=1):
        letter  = get_column_letter(col_idx)
        max_len = max(
            len(str(col)),
            df.iloc[:, col_idx - 1].astype(str).str.len().max() if len(df) > 0 else 0,
        )
        ws.column_dimensions[letter].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def export_report(report: dict, report_path: Path):
    """Exports a multi-sheet Excel report with summary and details."""

    summary_data = {
        "Metric": [
            "Source File",
            "Encoding Detected",
            "Original Rows",
            "Empty Rows Removed",
            "Duplicate Rows Removed",
            "Whitespace Cells Fixed",
            "Dates Standardized",
            "Column Names Fixed",
            "Final Row Count",
            "Cleaned File",
        ],
        "Value": [
            report["source_file"],
            report["encoding"],
            report["original_rows"],
            report["empty_removed"],
            report["duplicates_removed"],
            report["whitespace_fixed"],
            report["dates_fixed"],
            len(report["columns_renamed"]),
            report["final_rows"],
            report["output_file"],
        ],
    }

    summary_df = pd.DataFrame(summary_data)

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        if report["columns_renamed"]:
            col_df = pd.DataFrame(report["columns_renamed"], columns=["Original Name", "New Name"])
            col_df.to_excel(writer, sheet_name="Columns Renamed", index=False)

    wb = load_workbook(report_path)

    style_sheet(wb["Summary"], summary_df)

    # Color value column: green for good numbers, neutral for text
    ws = wb["Summary"]
    for row in range(2, len(summary_df) + 2):
        cell = ws.cell(row=row, column=2)
        val  = str(cell.value)
        if val.isdigit() and int(val) > 0 and row in [5, 6, 7, 8, 9]:
            cell.fill = PatternFill("solid", fgColor=GREEN_BG)

    if "Columns Renamed" in wb.sheetnames:
        col_df = pd.DataFrame(report["columns_renamed"], columns=["Original Name", "New Name"])
        style_sheet(wb["Columns Renamed"], col_df)

    wb.save(report_path)


# ── Core pipeline ──────────────────────────────────────────────────────────────

def clean_file(path: Path, output_dir: Path) -> dict:
    """Full cleaning pipeline for a single CSV file."""

    print(f"\n  📄 {path.name}")
    print(f"  {'─' * 36}")

    # Read
    df, encoding = fix_encoding(path)
    original_rows = len(df)
    print(f"  Encoding : {encoding}")
    print(f"  Rows     : {original_rows} original")

    # Clean
    df, cols_renamed    = clean_column_names(df)
    df, empty_removed   = remove_empty_rows(df)
    df, dupes_removed   = remove_duplicates(df)
    df, whitespace_fixed = trim_whitespace(df)
    df, dates_fixed     = standardize_dates(df)

    print(f"  ✓ Empty rows removed    : {empty_removed}")
    print(f"  ✓ Duplicates removed    : {dupes_removed}")
    print(f"  ✓ Whitespace cells fixed: {whitespace_fixed}")
    print(f"  ✓ Dates standardized    : {dates_fixed}")
    print(f"  ✓ Column names fixed    : {len(cols_renamed)}")
    print(f"  Final rows              : {len(df)}")

    # Save cleaned CSV
    output_csv  = output_dir / f"{path.stem}_cleaned.csv"
    report_path = output_dir / f"{path.stem}_report.xlsx"

    df.to_csv(output_csv, index=False, encoding="utf-8")

    # Build report
    report = {
        "source_file":        path.name,
        "encoding":           encoding,
        "original_rows":      original_rows,
        "empty_removed":      empty_removed,
        "duplicates_removed": dupes_removed,
        "whitespace_fixed":   whitespace_fixed,
        "dates_fixed":        dates_fixed,
        "columns_renamed":    cols_renamed,
        "final_rows":         len(df),
        "output_file":        output_csv.name,
    }

    export_report(report, report_path)

    print(f"\n  ✅ Cleaned : {output_csv.name}")
    print(f"  📊 Report  : {report_path.name}")

    return report


def find_csv_files(inputs: list[str]) -> list[Path]:
    """Resolves inputs to a list of CSV file paths."""
    files = []

    if len(inputs) == 0:
        files = sorted(Path(".").glob("*.csv"))

    elif len(inputs) == 1 and Path(inputs[0]).is_dir():
        files = sorted(Path(inputs[0]).glob("*.csv"))

    else:
        for f in inputs:
            p = Path(f)
            if not p.exists():
                print(f"  [WARN] Not found, skipping: {p}")
                continue
            if p.suffix.lower() != ".csv":
                print(f"  [WARN] Not a CSV file, skipping: {p}")
                continue
            files.append(p)

    return files


def main():
    parser = argparse.ArgumentParser(
        description="Clean messy CSV files and export an Excel report.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python clean_csv.py data.csv
  python clean_csv.py file1.csv file2.csv
  python clean_csv.py ./exports
  python clean_csv.py data.csv --output-dir ./results
        """,
    )
    parser.add_argument("inputs", nargs="*", help="CSV file(s) or folder")
    parser.add_argument(
        "--output-dir", "-d", default="output",
        help="Folder to save cleaned files and reports (default: output/)"
    )
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n🧹 CSV Cleaner Tool")
    print(f"{'─' * 40}")

    files = find_csv_files(args.inputs)

    if not files:
        print("  No CSV files found.")
        print("  Tips:")
        print("  • Make sure the folder contains .csv files")
        print("  • Or pass the file path directly: python clean_csv.py data.csv")
        sys.exit(0)

    print(f"  Found {len(files)} CSV file(s) to clean...")

    reports = []
    for f in files:
        r = clean_file(f, output_dir)
        reports.append(r)

    total_dupes  = sum(r["duplicates_removed"] for r in reports)
    total_empty  = sum(r["empty_removed"] for r in reports)
    total_ws     = sum(r["whitespace_fixed"] for r in reports)
    total_dates  = sum(r["dates_fixed"] for r in reports)

    print(f"\n{'─' * 40}")
    print(f"✅ All done! {len(files)} file(s) cleaned.")
    print(f"   Duplicates removed : {total_dupes}")
    print(f"   Empty rows removed : {total_empty}")
    print(f"   Whitespace fixed   : {total_ws}")
    print(f"   Dates standardized : {total_dates}")
    print(f"   Results saved to   : {output_dir}/")
    print(f"{'─' * 40}\n")


if __name__ == "__main__":
    main()
